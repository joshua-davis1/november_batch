import logging
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

months = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}

fileName = "expedia_report_monthly_march_2018.xlsx"

month = int(months[fileName.split("_")[-2]])
year = int(fileName.split("_")[-1][:4])
qDate = datetime.date(year,month,1)

recordList = []
headers = []

def logStart():
    logging.basicConfig(filename='scraper.log', level=logging.DEBUG)
    logging.debug('Analysing %s' % fileName.split(".")[0])  
# add first column to [headers], rest to [recordList]
def cleanMomRecords():
    for row in range(1,14):
        record = []
        for col in range(1,7):
            char = get_column_letter(col)
            if row == 1: headers.append(ws[char + str(row)].value)
            else: record.append(ws[char + str(row)].value)
        if len(record) != 0: recordList.append(record)
    headers[0] = 'Date'

def logRecord(fList):
    for record in fList:
        i = 0
        for field in record:
            logField(i,field)
            i+=1 

def cleanVocRecords():
    # add all records to lists
    for row in range(1,10):
        record = []
        for col in range(2,13):
            char = get_column_letter(col)
            record.append(ws[char + str(row)].value)
        if len(record) != 0: recordList.append(record)
    recordList.pop(1)
# convert all string reps to datetime objs
def cleanMonths():
    i = 0
    for field in recordList[0]:
        if type(field) == str:
            recordList[0][i] = datetime.date(year,months[field.lower()],1)
            i+=1

def addRecord(i):
    record = []
    record.append(recordList[0][i])
    record.append(recordList[1][i])
    record.append(recordList[2][i])
    record.append(recordList[3][i])
    record.append(recordList[4][i])
    record.append(recordList[5][i])
    record.append(recordList[6][i])
    record.append(recordList[7][i])
    return record

def transformVocRecords():
    return list(map(addRecord,range(0,9)))

def addVocReq():
    for record in recordList:
        if record[2] > 200: record[2] = "good"
    else: record[2] = "bad"
    if record[4] > 100: record[4] = "good"
    else: record[4] = "bad"
    if record[6] > 100: record[6] = "good"
    else: record[6] = "bad"

def printVocData():
    for row in filtered:
        for record in row:
            for field in record:
                print(field)
    
logStart()
try:
    wb = load_workbook(fileName)
    try:
        # task 1
        logging.debug("Starting first task")
        ws = wb['Summary Rolling MoM']
        cleanMomRecords()
        # filter list by requested (inputed) date
        filtered = filter(lambda record: record[0].date().month == qDate.month and record[0].date().year == qDate.year, recordList)
        logField = lambda i, field :logging.info("%s: %s" % (headers[i],field))
        logRecord(filtered)
        logging.debug("Succesfully completed first task.")
    except:
        logging.error("An error occurred in first task.")
    try:
        # task 2
        logging.debug("Starting second task")
        recordList = []
        headers = []
        ws = wb['VOC Rolling MoM']
        cleanVocRecords()
        cleanMonths()
        recordList = transformVocRecords()
        addVocReq()
        filtered = [filter(lambda record: record[0].month == qDate.month and record[0].year == qDate.year, recordList)]
        printVocData()
        logging.debug("Succesfully completed second task.")
    except:
        logging.error("An error occurred in our second task.")

except OSError as error:
    logging.error("Failed to load %s\n", fileName, error)
finally:
    wb.close()